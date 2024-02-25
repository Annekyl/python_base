# win+R输入cmd进入命令提示符，输入pip install tabulate下载模块
from random import randint
import openpyxl

print("--------------超市商品录入与查询系统----------------")
system = openpyxl.load_workbook(u"超市管理系统.xlsx")
sheet1 = system["账户"]
sheet2 = system["商品统计"]
sheet3 = system["销售额"]
point = 0
while point == 0:
    sel = input("1.登录 2.注册 3.退出")
    if sel == "1":
        user = int(input("请输入员工码"))
        password = int(input("请输入密码"))
        cheek = randint(1000, 9999)
        is_cheek = int(input(f"请输入验证码{cheek}"))
        sheet1_line_2 = sheet1["B"]
        sheet1_line_3 = sheet1["C"]
        num = len(sheet1_line_2)
        for i in range(num):
            is_user = sheet1_line_2[i].value
            is_password = sheet1_line_3[i].value
            if user == is_user and password == is_password and cheek == is_cheek:
                print("登录成功！")
                point = 1
                break
        else:
            print("登录失败！")
    if sel == "2":
        user = int(input("请输入员工码"))
        password = int(input("请输入密码"))
        cheek = randint(1000, 9999)
        is_cheek = int(input(f"请输入验证码{cheek}"))
        sheet1_line_2 = sheet1["B"]
        sheet1_line_3 = sheet1["C"]
        num = len(sheet1_line_2)
        if is_cheek == cheek:
            sheet1.cell(num + 1, 1).value = num
            sheet1.cell(num + 1, 2).value = user
            sheet1.cell(num + 1, 3).value = password
            system.save(u"超市管理系统.xlsx")
            print("注册成功！")
        else:
            print("注册失败！")
    if sel == "3":
        break
while point == 1:
    a = int(input(
        "----------主菜单----------\n1.查询超市全部商品\n2.商品录入\n3.单个商品查询\n4.收银\n5.查看销售额 6.退出\n"))
    if a == 1:
        sheet2_line_1 = sheet2["A"]
        sheet2_line_2 = sheet2["B"]
        sheet2_line_3 = sheet2["C"]
        sheet2_line_4 = sheet2["D"]
        sheet2_line_5 = sheet2["E"]
        num = len(sheet2_line_1)
        for i in range(num):
            print(sheet2_line_1[i].value, end="\t")
            print(sheet2_line_2[i].value, end="\t")
            print(sheet2_line_3[i].value, end="\t")
            print(sheet2_line_4[i].value, end="\t")
            print(sheet2_line_5[i].value)
    if a == 2:
        new_id = int(input("请扫描商品条形码"))
        new_name = input("请输入商品名称")
        new_price = float(input("请输入商品价格"))
        new_number = int(input("请输入商品数量"))
        print("商品已录入成功")
        sheet2_line_1 = sheet2["A"]
        num = len(sheet2_line_1)
        sheet2.cell(num + 1, 1).value = num
        sheet2.cell(num + 1, 2).value = new_id
        sheet2.cell(num + 1, 3).value = new_name
        sheet2.cell(num + 1, 4).value = new_price
        sheet2.cell(num + 1, 5).value = new_number
        system.save(u"超市管理系统.xlsx")
    if a == 3:
        cheek_id = int(input("请扫描商品条形码"))
        sheet2_line_1 = sheet2["A"]
        sheet2_line_2 = sheet2["B"]
        sheet2_line_3 = sheet2["C"]
        sheet2_line_4 = sheet2["D"]
        sheet2_line_5 = sheet2["E"]
        num = len(sheet2_line_1)
        for i in range(num):
            if cheek_id == sheet2_line_2[i].value:
                print(sheet2_line_1[i].value, end="\t")
                print(sheet2_line_2[i].value, end="\t")
                print(sheet2_line_3[i].value, end="\t")
                print(sheet2_line_4[i].value, end="\t")
                print(sheet2_line_5[i].value)
                break
        else:
            print("商品不存在！")
    if a == 4:
        is_sale = 1
        while is_sale == 1:
            sheet2_line_1 = sheet2["A"]
            sheet2_line_2 = sheet2["B"]
            sheet2_line_3 = sheet2["C"]
            sheet2_line_4 = sheet2["D"]
            sheet2_line_5 = sheet2["E"]
            sheet2_num = len(sheet2_line_1)
            sheet3_line_1 = sheet3["A"]
            sheet3_line_2 = sheet3["B"]
            sheet3_line_3 = sheet3["C"]
            sheet3_line_4 = sheet3["D"]
            sheet3_line_5 = sheet3["E"]
            sheet3_num = len(sheet3_line_1)
            sale_id = int(input("请扫描商品条形码"))
            for i in range(sheet2_num):  # 查找商品信息
                if sheet2_line_2[i].value == sale_id:  # 存在商品信息
                    sheet2.cell(i + 1, 5).value -= 1
                    for j in range(sheet3_num):  # 查找是否卖过同类商品
                        if sheet3_line_2[j] == sale_id:  # 如果存在，销售数量加1
                            sheet3.cell(j, 5).value += 1  # 销售数量加1
                            system.save(u"超市管理系统.xlsx")
                            print("出库成功")
                            break
                    else:  # 如果不存在，追加到sheet3
                        num = 0
                        for i in range(sheet2_num):
                            num += 1
                            if sheet2_line_2[i] == sale_id:
                                break
                        sheet3.cell(sheet3_num + 1, 1).value = sheet2.cell(num, 1).value
                        sheet3.cell(sheet3_num + 1, 2).value = sheet2.cell(num, 2).value
                        sheet3.cell(sheet3_num + 1, 3).value = sheet2.cell(num, 3).value
                        sheet3.cell(sheet3_num + 1, 4).value = sheet2.cell(num, 4).value
                        sheet3.cell(sheet3_num + 1, 5).value = 1
                        system.save(u"超市管理系统.xlsx")
                        print("出库成功")
                        break
            else:
                print("商品不存在，请先录入商品")
            is_sale = int(input("请选择 1.继续扫描 2.结束收银"))

        receive = input("请扫描收款码")
        is_correct = receive[0:2:1]
        is_correct = int(is_correct)
        if is_correct == 10 or is_correct == 11 or is_correct == 12 or is_correct == 13 \
                or is_correct == 14 or is_correct == 15:
            if len(receive) == 18:
                print("付款成功，谢谢惠顾！")
            else:
                print("付款失败！")
        else:
            print("付款失败！")
    if a == 5:
        sheet3_line_1 = sheet3["A"]
        sheet3_line_2 = sheet3["B"]
        sheet3_line_3 = sheet3["C"]
        sheet3_line_4 = sheet3["D"]
        sheet3_line_5 = sheet3["E"]
        num = len(sheet3_line_1)
        for i in range(num):
            print(sheet3_line_1[i].value, end="\t")
            print(sheet3_line_2[i].value, end="\t")
            print(sheet3_line_3[i].value, end="\t")
            print(sheet3_line_4[i].value, end="\t")
            print(sheet3_line_5[i].value)
    if a == 6:
        print("超市商品录入与查询系统已退出！")
        break
system.save(u"超市管理系统.xlsx")
system.close()
