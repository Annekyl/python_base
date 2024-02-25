import openpyxl
import os


# 读取文本文件内容
def read_text_file(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return content


# 将文本内容按字节"翻译"到Excel表
def write_to_excel(content, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active

    for i, byte in enumerate(content):
        row = i // 256 + 1
        col = i % 256 + 1
        ws.cell(row=row, column=col).value = byte

    wb.save(output_file)


# 将Excel表中的数据解读成字符和二进制
def read_from_excel(input_file):
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    content = ''
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                content += chr(cell.value)

    return content


# 主函数
def main():
    input_file = 'input.txt'
    output_file = 'output.xlsx'

    content = read_text_file(input_file)
    write_to_excel(content, output_file)
    translated_content = read_from_excel(output_file)

    print('原文本：', content)
    print('翻译后的文本：', translated_content)


if __name__ == '__main__':
    main()
