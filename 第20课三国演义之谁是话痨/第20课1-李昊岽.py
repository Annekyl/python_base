import openpyxl
from openpyxl import load_workbook

workbook = openpyxl.Workbook()
worksheet = workbook.active

with open('1.txt', 'r', encoding='utf-8') as f:
    column,n,row,biao= 1,1,1,'A'
    for line in f:
        for char in line:
            a = ord(char)
            b = chr(ord(biao)+len(char.encode('utf-8')))
            #print(biao,b)
            worksheet.cell(row=row,column=column,value=n)
            worksheet.merge_cells(f'{biao}{row}:{b}{row}')
            worksheet.cell(row=row+1,column=column,value=char)
            worksheet.merge_cells(f'{biao}{row+1}:{b}{row+1}')
            worksheet.cell(row=row + 2, column=column,value=a)
            worksheet.merge_cells(f'{biao}{row+2}:{b}{row+2}')
            worksheet.cell(row=row + 3, column=column,value=bytes(char,'utf-8').hex())
            worksheet.merge_cells(f'{biao}{row+3}:{b}{row+3}')            
            biao=chr(ord(biao)+len(char.encode('utf-8'))+1)
            column += len(char.encode('utf-8'))+1
            n+=1
        row += 4  
        column ,biao= 1,'A'
workbook.save('example.xlsx')
