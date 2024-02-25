# ----------------载入库---------------
import openpyxl

# ---------------主程序--------------------
fr = open("任务1.txt", "r", encoding="UTF-8")
fw = openpyxl.load_workbook("任务1.xlsx")
sheet = fw["Sheet1"]
num = 1  # 序号
cell_row = 1  # 单元格的行
cell_line = 1  # 单元格的列
begin = "A"
for line in fr:  # 依次读取每一列
    for i in line:  # 依次读取每一个字符
        a = ord(i)  # UTF-8编码
        hebing = chr(ord(begin) + len(i.encode("UTF-8")))
        # 填序号
        sheet.cell(cell_row, cell_line).value = num
        sheet.merge_cells(f"{begin}{cell_row}:{hebing}{cell_row}")
        # 合并单元格并填入原内容
        sheet.cell(cell_row + 1, cell_line).value = i
        sheet.merge_cells(f"{begin}{cell_row + 1}:{hebing}{cell_row + 1}")
        # 合并单元格并填入UTF-8编码
        sheet.cell(cell_row + 2, cell_line).value = a
        sheet.merge_cells(f"{begin}{cell_row + 2}:{hebing}{cell_row + 2}")
        sheet.cell(cell_row + 3, cell_line).value = bytes(i, "UTF-8").hex()
        sheet.merge_cells(f"{begin}{cell_row + 3}:{hebing}{cell_row + 3}")
        # 为下次循环做准备
        num += 1
        cell_line += len(i.encode("UTF-8")) + 1
    cell_row += 4
    cell_line = 1
    begin = "A"
# 关闭文件
fr.close()
fw.save("任务1.xlsx")
fw.close()
